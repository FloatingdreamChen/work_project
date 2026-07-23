"""Import civil service position tables into PostgreSQL.

CSV is supported with the Python standard library. XLSX is supported when
openpyxl is available in the active environment.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
from datetime import datetime
from pathlib import Path

from sqlalchemy import text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import gov exam positions.")
    parser.add_argument("file", type=Path, help="CSV or XLSX position table.")
    parser.add_argument("--exam-year", type=int, required=True)
    parser.add_argument("--exam-type", required=True, help="国考、省考、事业单位等")
    return parser.parse_args()


FIELD_ALIASES = {
    "province": ("省份", "地区", "province"),
    "city": ("城市", "city"),
    "department": ("部门", "招录机关", "department"),
    "bureau": ("用人司局", "机构", "bureau"),
    "position_name": ("职位名称", "岗位名称", "position_name"),
    "position_code": ("职位代码", "岗位代码", "position_code"),
    "recruitment_count": ("招录人数", "人数", "recruitment_count"),
    "applicant_count": ("报名人数", "缴费人数", "applicant_count"),
    "competition_ratio": ("竞争比", "竞争比例", "competition_ratio"),
    "previous_min_score": ("往年最低分", "进面最低分", "previous_min_score"),
    "education_requirement": ("学历", "学历要求", "education_requirement"),
    "degree_requirement": ("学位", "学位要求", "degree_requirement"),
    "major_requirement": ("专业", "专业要求", "major_requirement"),
    "political_requirement": ("政治面貌", "political_requirement"),
    "grassroots_requirement": ("基层工作最低年限", "基层经历", "grassroots_requirement"),
    "work_years_requirement": ("工作年限", "work_years_requirement"),
    "household_requirement": ("户籍", "生源地", "household_requirement"),
    "remarks": ("备注", "remarks"),
    "source_name": ("来源", "source_name"),
    "source_url": ("来源链接", "source_url"),
    "source_published_at": ("发布日期", "发布时间", "source_published_at"),
}


def get_value(row: dict, field: str):
    for key in FIELD_ALIASES[field]:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def read_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fp:
            return list(csv.DictReader(fp))
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("XLSX import requires openpyxl. Use CSV or install openpyxl.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [dict(zip(headers, values)) for values in rows[1:]]
    raise SystemExit("Only CSV and XLSX files are supported.")


async def import_rows(rows: list[dict], exam_year: int, exam_type: str, file_path: str = "") -> int:
    from backend.db.session import AsyncSessionLocal

    fields = ["exam_year", "exam_type", *FIELD_ALIASES.keys()]
    tail = {"source_name", "source_url", "source_published_at"}
    fields = [field for field in fields if field not in tail] + ["source_name", "source_url", "source_published_at"]
    columns = ", ".join(fields)
    params = ", ".join([f":{field}" for field in fields])
    count = 0
    errors = []
    async with AsyncSessionLocal() as db:
        for row_index, row in enumerate(rows, start=2):
            payload = {field: get_value(row, field) for field in FIELD_ALIASES}
            payload["exam_year"] = exam_year
            payload["exam_type"] = exam_type
            payload["position_name"] = payload.get("position_name") or "未命名岗位"
            for int_field in ("recruitment_count", "applicant_count"):
                payload[int_field] = _to_int(payload.get(int_field))
            for float_field in ("competition_ratio", "previous_min_score"):
                payload[float_field] = _to_float(payload.get(float_field))
            payload["source_published_at"] = _to_datetime_iso(payload.get("source_published_at"))
            try:
                await db.execute(
                    text(f"INSERT INTO positions ({columns}) VALUES ({params})"),
                    payload,
                )
                count += 1
            except Exception as exc:
                errors.append({"row": row_index, "error": str(exc), "position_name": payload.get("position_name")})
        await db.execute(
            text(
                """
                INSERT INTO position_import_audits (
                    source_name, file_path, total_rows, imported_rows, failed_rows, errors
                )
                VALUES (:source_name, :file_path, :total_rows, :imported_rows, :failed_rows, CAST(:errors AS JSONB))
                """
            ),
            {
                "source_name": get_value(rows[0], "source_name") if rows else None,
                "file_path": file_path,
                "total_rows": len(rows),
                "imported_rows": count,
                "failed_rows": len(errors),
                "errors": json.dumps(errors, ensure_ascii=False),
            },
        )
        await db.commit()
    return count


def _to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except ValueError:
        return None


def _to_float(value):
    if value in (None, ""):
        return None
    text_value = str(value).replace(":1", "").replace("：1", "").replace(",", "").strip()
    try:
        return float(text_value)
    except ValueError:
        return None


def _to_datetime_iso(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    text_value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text_value, fmt).isoformat()
        except ValueError:
            continue
    return text_value


async def amain() -> None:
    args = parse_args()
    if not args.file.exists():
        raise SystemExit(f"File not found: {args.file}")
    rows = read_rows(args.file)
    count = await import_rows(rows, args.exam_year, args.exam_type, str(args.file.resolve()))
    print(f"Positions imported: {count}")


if __name__ == "__main__":
    asyncio.run(amain())
